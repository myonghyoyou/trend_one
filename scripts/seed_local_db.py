"""Load the documented XLSX fixtures into the local PostgreSQL container."""

import argparse
import csv
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_CONTAINER = "trend-one-postgres"


def format_value(value):
    if value is None or value == "":
        return r"\N"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def write_csv(path, header, rows):
    with path.open("w", encoding="utf-8", newline="") as output:
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(header)
        writer.writerows([[format_value(value) for value in row] for row in rows])


def read_sheet(path, header_row=0, data_start_row=1):
    worksheet = load_workbook(path, read_only=True, data_only=True).active
    rows = worksheet.iter_rows(values_only=True)
    for _ in range(header_row):
        next(rows)
    header = tuple(next(rows))
    data = []
    for index, row in enumerate(rows):
        if index + header_row + 1 < data_start_row:
            continue
        row = tuple(row)
        if any(value is not None and value != "" for value in row):
            data.append(row)
    return header, data


def run(docker, *arguments):
    command = [docker, *arguments]
    subprocess.run(command, check=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--docker", default="docker")
    parser.add_argument("--container", default=DEFAULT_CONTAINER)
    args = parser.parse_args()

    governor_path = ROOT / "docs" / "t_governor_202607161800.xlsx"
    stat_path = ROOT / "docs" / "t_governor_stat_202607161800.xlsx"
    region_path = ROOT / "docs" / "t_region_cd_202607161759.xlsx"

    governor_header, governor_rows = read_sheet(governor_path)
    stat_header, stat_rows = read_sheet(stat_path)
    region_header, region_rows = read_sheet(region_path)

    with tempfile.TemporaryDirectory(prefix="trend-one-seed-") as temp_dir:
        temp = Path(temp_dir)
        governor_csv = temp / "t_governor.csv"
        stat_csv = temp / "t_governor_stat.csv"
        region_csv = temp / "t_region_cd.csv"

        write_csv(governor_csv, governor_header, governor_rows)
        write_csv(stat_csv, stat_header, stat_rows)
        write_csv(region_csv, region_header, region_rows)

        run(
            args.docker,
            "exec",
            args.container,
            "psql",
            "-v",
            "ON_ERROR_STOP=1",
            "-U",
            "trend_one",
            "-d",
            "trend_one_dev",
            "-c",
            "TRUNCATE TABLE t_governor_stat, t_governor, t_region_cd RESTART IDENTITY;",
        )

        for csv_path in (governor_csv, stat_csv, region_csv):
            run(args.docker, "cp", str(csv_path), f"{args.container}:/tmp/{csv_path.name}")

        copy_commands = {
            "t_governor.csv": "COPY t_governor (gvrnr_uid, gvrnr_nm, cate_cd, rgst_dttm, rgst_uid, updt_dttm, updt_uid, inspct_day) FROM '/tmp/t_governor.csv' WITH (FORMAT csv, HEADER true, NULL '\\N');",
            "t_governor_stat.csv": "COPY t_governor_stat (gvrnr_uid, gvrnr_press1, gvrnr_press2, gvrnr_trnsps1, gvrnr_trnsps2, record_dttm, rgst_dttm, rgst_uid, updt_dttm, updt_uid) FROM '/tmp/t_governor_stat.csv' WITH (FORMAT csv, HEADER true, NULL '\\N');",
            "t_region_cd.csv": "COPY t_region_cd (lvl, up_cd, cate_cd, cd_name) FROM '/tmp/t_region_cd.csv' WITH (FORMAT csv, HEADER true, NULL '\\N');",
        }
        for csv_name in ("t_region_cd.csv", "t_governor.csv", "t_governor_stat.csv"):
            run(
                args.docker,
                "exec",
                args.container,
                "psql",
                "-v",
                "ON_ERROR_STOP=1",
                "-U",
                "trend_one",
                "-d",
                "trend_one_dev",
                "-c",
                copy_commands[csv_name],
            )

        run(
            args.docker,
            "exec",
            args.container,
            "psql",
            "-U",
            "trend_one",
            "-d",
            "trend_one_dev",
            "-c",
            "SELECT 't_region_cd' AS table_name, COUNT(*) AS row_count FROM t_region_cd UNION ALL SELECT 't_governor', COUNT(*) FROM t_governor UNION ALL SELECT 't_governor_stat', COUNT(*) FROM t_governor_stat ORDER BY table_name;",
        )


if __name__ == "__main__":
    main()
